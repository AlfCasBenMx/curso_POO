"""
Generador de datasets financieros para Sesión 1 — POC Thinking con Copilot
Genera:
  1. gastos_departamentales.xlsx  (~315 filas) — presupuesto vs gasto real por departamento
  2. cartera_clientes.xlsx        (~80 filas)  — cuentas por cobrar con aging

Ambos datasets están diseñados para:
  ✅ Ejercicios con Copilot en Excel (fórmulas, tablas dinámicas, gráficos)
  ✅ Carga en Power BI (dashboard, Q&A en lenguaje natural)
  ✅ Detección de anomalías y datos sucios (duplicados, outliers)
  ✅ Datos ya formateados como Tabla de Excel (ListObject) — Copilot los reconoce directamente
"""

import random
import os
from datetime import datetime, timedelta

try:
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ Necesitas instalar openpyxl: pip install openpyxl")
    exit(1)

random.seed(42)

OUT_DIR = os.path.dirname(os.path.abspath(__file__))

# ─────────────────────────────────────────────
# 1. GASTOS DEPARTAMENTALES
# ─────────────────────────────────────────────

departamentos = {
    "Finanzas":    {"cc": "CC-100", "responsables": ["Ana López", "Carlos Mendoza"]},
    "Operaciones": {"cc": "CC-200", "responsables": ["Ricardo Torres", "Sofía Hernández"]},
    "TI":          {"cc": "CC-300", "responsables": ["Miguel Ángel Ruiz", "Laura Castillo"]},
    "RH":          {"cc": "CC-400", "responsables": ["Patricia Díaz", "Javier Morales"]},
    "Legal":       {"cc": "CC-500", "responsables": ["Fernanda Reyes"]},
    "Comercial":   {"cc": "CC-600", "responsables": ["Andrés García", "Daniela Vargas"]},
}

categorias_gasto = [
    ("Nómina", (150_000, 500_000)),
    ("Servicios profesionales", (30_000, 200_000)),
    ("Software y licencias", (10_000, 80_000)),
    ("Viáticos y viajes", (5_000, 60_000)),
    ("Capacitación", (3_000, 40_000)),
    ("Material de oficina", (1_000, 15_000)),
    ("Renta y mantenimiento", (20_000, 120_000)),
    ("Publicidad y marketing", (10_000, 90_000)),
]

meses = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
         "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

rows_gastos = []
row_id = 1

for depto, info in departamentos.items():
    # Cada depto tiene 3-4 categorías de gasto por mes (no todas aplican a todos)
    cats_depto = random.sample(categorias_gasto, k=random.randint(3, 5))
    for mes_idx, mes in enumerate(meses):
        for cat_nombre, (low, high) in cats_depto:
            presupuesto = round(random.uniform(low, high), 2)
            # El gasto real varía ±25% del presupuesto, con algunos outliers
            variacion = random.gauss(1.0, 0.12)
            if random.random() < 0.05:  # 5% outliers extremos
                variacion = random.choice([1.5, 1.8, 0.3, 2.1])
            gasto_real = round(presupuesto * variacion, 2)
            responsable = random.choice(info["responsables"])
            estatus = random.choices(
                ["Aprobado", "Pendiente", "Rechazado"],
                weights=[75, 20, 5]
            )[0]

            rows_gastos.append({
                "ID": row_id,
                "Departamento": depto,
                "Centro_Costo": info["cc"],
                "Categoria_Gasto": cat_nombre,
                "Mes": mes,
                "Anio": 2025,
                "Presupuesto_MXN": presupuesto,
                "Gasto_Real_MXN": gasto_real,
                "Variacion_Pct": round((variacion - 1) * 100, 1),
                "Responsable": responsable,
                "Estatus": estatus,
            })
            row_id += 1

# Insertar 3 duplicados intencionales (para ejercicio de detección)
for dup in random.sample(rows_gastos[:60], 3):
    dup_copy = dict(dup)
    dup_copy["ID"] = row_id
    rows_gastos.append(dup_copy)
    row_id += 1

random.shuffle(rows_gastos)

# --- Escribir como .xlsx con formato de Tabla ---
def write_xlsx_with_table(filepath, rows, table_name, sheet_name="Datos"):
    """Escribe una lista de dicts como .xlsx con formato de Tabla de Excel (ListObject)."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Encabezados
    headers = list(rows[0].keys())
    ws.append(headers)

    # Datos
    for row in rows:
        ws.append([row[h] for h in headers])

    # Crear Tabla de Excel (equivalente a Ctrl+T)
    last_col = get_column_letter(len(headers))
    last_row = len(rows) + 1  # +1 por encabezado
    table_ref = f"A1:{last_col}{last_row}"

    tab = Table(displayName=table_name, ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Ajustar ancho de columnas
    for i, header in enumerate(headers, 1):
        col_letter = get_column_letter(i)
        max_len = max(len(str(header)), 12)
        ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(filepath)

gastos_path = os.path.join(OUT_DIR, "gastos_departamentales.xlsx")
write_xlsx_with_table(gastos_path, rows_gastos, "GastosDepartamentales")

print(f"✅ gastos_departamentales.xlsx → {len(rows_gastos)} filas (con formato Tabla)")


# ─────────────────────────────────────────────
# 2. CARTERA DE CLIENTES (Cuentas por cobrar)
# ─────────────────────────────────────────────

clientes = [
    "Grupo Alfa SA", "Industrias Beta", "Comercial Gamma", "Tech Delta",
    "Servicios Epsilon", "Logistics Zeta", "Constructora Eta",
    "Retail Theta", "Agro Iota", "Pharma Kappa", "Mining Lambda",
    "Telecom Mu", "Energy Nu", "Food Xi", "Auto Omicron",
]

regiones = ["Norte", "Centro", "Sur", "Occidente", "Noreste"]

rows_cartera = []
factura_num = 5001

base_date = datetime(2025, 6, 1)

for i in range(80):
    cliente = random.choice(clientes)
    region = random.choice(regiones)
    monto = round(random.uniform(15_000, 800_000), 2)

    # Fecha de emisión entre 1 y 180 días atrás desde base_date
    dias_atras = random.randint(1, 180)
    fecha_emision = base_date - timedelta(days=dias_atras)
    # Plazo de pago: 30, 60 o 90 días
    plazo = random.choice([30, 60, 90])
    fecha_vencimiento = fecha_emision + timedelta(days=plazo)
    dias_vencido = max(0, (base_date - fecha_vencimiento).days)

    if dias_vencido == 0:
        estatus = random.choices(["Pagada", "Vigente"], weights=[40, 60])[0]
    elif dias_vencido <= 30:
        estatus = random.choices(["Pagada", "Vencida 1-30"], weights=[30, 70])[0]
    elif dias_vencido <= 60:
        estatus = random.choices(["Pagada", "Vencida 31-60"], weights=[20, 80])[0]
    elif dias_vencido <= 90:
        estatus = "Vencida 61-90"
    else:
        estatus = "Vencida 90+"

    if dias_vencido <= 30:
        aging = "Corriente"
    elif dias_vencido <= 60:
        aging = "30-60 días"
    elif dias_vencido <= 90:
        aging = "60-90 días"
    else:
        aging = "90+ días"

    # Riesgo estimado
    if "Pagada" in estatus:
        riesgo = "Bajo"
        provision_pct = 0
    elif aging == "Corriente":
        riesgo = "Bajo"
        provision_pct = 1
    elif aging == "30-60 días":
        riesgo = "Medio"
        provision_pct = 5
    elif aging == "60-90 días":
        riesgo = "Alto"
        provision_pct = 15
    else:
        riesgo = "Crítico"
        provision_pct = 40

    rows_cartera.append({
        "ID_Factura": f"FAC-{factura_num}",
        "Cliente": cliente,
        "Region": region,
        "Monto_MXN": monto,
        "Fecha_Emision": fecha_emision.strftime("%Y-%m-%d"),
        "Fecha_Vencimiento": fecha_vencimiento.strftime("%Y-%m-%d"),
        "Plazo_Dias": plazo,
        "Dias_Vencido": dias_vencido,
        "Estatus_Pago": estatus,
        "Categoria_Aging": aging,
        "Riesgo": riesgo,
        "Provision_Pct": provision_pct,
        "Provision_MXN": round(monto * provision_pct / 100, 2),
    })
    factura_num += 1

# Insertar 2 posibles duplicados (misma factura, mismo monto — error de captura)
for dup in random.sample(rows_cartera[:20], 2):
    dup_copy = dict(dup)
    dup_copy["ID_Factura"] = f"FAC-{factura_num}"  # diferente ID pero mismo cliente/monto
    rows_cartera.append(dup_copy)
    factura_num += 1

random.shuffle(rows_cartera)

cartera_path = os.path.join(OUT_DIR, "cartera_clientes.xlsx")
write_xlsx_with_table(cartera_path, rows_cartera, "CarteraClientes")

print(f"✅ cartera_clientes.xlsx       → {len(rows_cartera)} filas (con formato Tabla)")
print(f"\n📂 Archivos generados en: {OUT_DIR}")
print("\n✨ Los archivos .xlsx ya tienen formato de Tabla (Ctrl+T).")
print("   Solo necesitas subirlos a OneDrive y abrirlos — Copilot los reconoce directamente.")
print("\nPróximos pasos:")
print("  1. Sube los .xlsx a OneDrive/SharePoint")
print("  2. Abre en Excel → verifica que Autoguardado esté ON")
print("  3. Verifica que aparece el ícono de Copilot")
print("  4. ¡Listo para usar!")
