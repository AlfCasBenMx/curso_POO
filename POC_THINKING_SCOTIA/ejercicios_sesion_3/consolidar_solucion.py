"""
Solución completa — Sesión 3: Consolidación de múltiples fuentes → Formato Global

Procesa las 5 fuentes de datos_entrada/:
  banca_comercial (150), banca_corporativa (120), seguros (100),
  tarjetas (130), tesoreria (90).  Total = 590 filas.

Este script:
1. Lee las 5 fuentes Excel de datos_entrada/
2. Homologa nombres de columnas
3. Clasifica registros (Ingreso / Gasto / Provisión)
4. Genera 3 Excels consolidados
5. Genera el Formato Global por línea de negocio
6. Genera log de excepciones y validaciones
"""

import os
import glob
import pandas as pd
import warnings
warnings.filterwarnings("ignore")

# ═══════════════════════════════════════════════════════════
# CONFIGURACIÓN
# ═══════════════════════════════════════════════════════════

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR = os.path.join(SCRIPT_DIR, "datos_entrada")
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "datos_salida")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ═══════════════════════════════════════════════════════════
# PASO 1: LEER TODAS LAS FUENTES
# ═══════════════════════════════════════════════════════════

print("=" * 60)
print("📥 PASO 1: Leyendo fuentes de datos")
print("=" * 60)

# Diccionario para almacenar todos los DataFrames
dataframes = {}

# Leer archivos xlsx
archivos_xlsx = glob.glob(os.path.join(INPUT_DIR, "*.xlsx"))
for archivo in sorted(archivos_xlsx):
    nombre = os.path.basename(archivo)
    df = pd.read_excel(archivo)
    df["archivo_origen"] = nombre
    dataframes[nombre] = df
    print(f"  📁 {nombre} — {len(df)} filas × {len(df.columns)} columnas")
    print(f"     Columnas: {', '.join(df.columns[:6])}...")

total_archivos = len(dataframes)
total_registros_raw = sum(len(df) for df in dataframes.values())
print(f"\n  ✅ {total_archivos} archivos leídos — {total_registros_raw} registros totales")

# ═══════════════════════════════════════════════════════════
# PASO 2: HOMOLOGAR COLUMNAS
# ═══════════════════════════════════════════════════════════

print("\n" + "=" * 60)
print("🔀 PASO 2: Homologando columnas")
print("=" * 60)

# Mapeo de columnas por archivo fuente → columnas estándar
MAPEO_COLUMNAS = {
    "banca_comercial.xlsx": {
        "fecha_registro": "fecha",
        "linea_negocio": "linea_negocio",
        "concepto_operacion": "concepto",
        "subcuenta_contable": "subcuenta",
        "monto_MXN": "monto",
        "tipo_movimiento": "tipo",
        "periodo": "periodo",
        "responsable": "responsable",
    },
    "banca_corporativa.xlsx": {
        "date": "fecha",
        "business_line": "linea_negocio",
        "concept": "concepto",
        "sub_account": "subcuenta",
        "amount_MXN": "monto",
        "type": "tipo",
        "period": "periodo",
        "owner": "responsable",
    },
    "seguros.xlsx": {
        "fecha": "fecha",
        "area": "linea_negocio",
        "tipo_operacion": "concepto",
        "cta_subcuenta": "subcuenta",
        "importe": "monto",
        "naturaleza": "tipo",
        "mes": "periodo",
        "responsable_area": "responsable",
    },
    "tarjetas.xlsx": {
        "fecha_movimiento": "fecha",
        "producto": "linea_negocio",
        "concepto": "concepto",
        "num_cuenta": "subcuenta",
        "monto_pesos": "monto",
        "clasificacion": "tipo",
        "periodo_contable": "periodo",
        "ejecutivo": "responsable",
    },
    "tesoreria.xlsx": {
        "value_date": "fecha",
        "treasury_desk": "linea_negocio",
        "transaction_type": "concepto",
        "account_code": "subcuenta",
        "amount_MXN": "monto",
        "pnl_type": "tipo",
        "month": "periodo",
        "dealer": "responsable",
    },
}

COLUMNAS_ESTANDAR = ["fecha", "linea_negocio", "concepto", "subcuenta",
                     "monto", "tipo", "periodo", "responsable", "archivo_origen"]

# Aplicar mapeo y concatenar
dfs_homologados = []
for nombre_archivo, df in dataframes.items():
    mapeo = MAPEO_COLUMNAS.get(nombre_archivo, {})
    if mapeo:
        df_mapped = df.rename(columns=mapeo)
        # Seleccionar solo las columnas estándar que existan
        cols_disponibles = [c for c in COLUMNAS_ESTANDAR if c in df_mapped.columns]
        df_mapped = df_mapped[cols_disponibles]
        dfs_homologados.append(df_mapped)
        print(f"  ✅ {nombre_archivo} — homologado ({len(cols_disponibles)} columnas)")
    else:
        print(f"  ⚠️ {nombre_archivo} — sin mapeo definido, omitido")

# Concatenar todos
df_consolidado = pd.concat(dfs_homologados, ignore_index=True)
print(f"\n  📊 DataFrame consolidado: {len(df_consolidado)} filas × {len(df_consolidado.columns)} columnas")

# ═══════════════════════════════════════════════════════════
# PASO 3: CLASIFICAR REGISTROS
# ═══════════════════════════════════════════════════════════

print("\n" + "=" * 60)
print("🏷️ PASO 3: Clasificando registros")
print("=" * 60)

def clasificar(concepto):
    """Clasifica un concepto como Ingreso, Gasto, Provisión o Sin clasificar."""
    if pd.isna(concepto) or concepto == "":
        return "Sin clasificar"
    
    concepto_lower = str(concepto).lower()
    
    # Ingresos
    if any(kw in concepto_lower for kw in [
        "comisión", "comision", "spread", "prima cobrada", "prima neta",
        "rendimiento", "interés", "interes", "asesoría", "asesoria",
        "anualidad", "intereses tarjeta"
    ]):
        return "Ingreso"
    
    # Gastos
    if any(kw in concepto_lower for kw in [
        "gasto", "nómina", "nomina", "personal", "servicio", "licencia",
        "renta", "mantenimiento", "multa", "publicidad", "seguridad",
        "ti externo", "limpieza"
    ]):
        return "Gasto"
    
    # Provisiones
    if any(kw in concepto_lower for kw in [
        "reserva", "pérdida", "perdida", "provisión", "provision",
        "valuación", "valuacion"
    ]):
        return "Provisión"
    
    return "Sin clasificar"

df_consolidado["clasificacion"] = df_consolidado["concepto"].apply(clasificar)

# Conteos
conteos = df_consolidado["clasificacion"].value_counts()
for cat, count in conteos.items():
    emoji = {"Ingreso": "💵", "Gasto": "💸", "Provisión": "📦", "Sin clasificar": "⚠️"}.get(cat, "❓")
    print(f"  {emoji} {cat}: {count} registros")

# ═══════════════════════════════════════════════════════════
# PASO 4: GENERAR 3 EXCELS CONSOLIDADOS
# ═══════════════════════════════════════════════════════════

print("\n" + "=" * 60)
print("📤 PASO 4: Generando Excels consolidados")
print("=" * 60)

for clasificacion, nombre_archivo in [
    ("Ingreso", "consolidado_ingresos.xlsx"),
    ("Gasto", "consolidado_gastos.xlsx"),
    ("Provisión", "consolidado_provisiones.xlsx"),
]:
    df_filtrado = df_consolidado[df_consolidado["clasificacion"] == clasificacion].copy()
    
    if len(df_filtrado) == 0:
        print(f"  ⚠️ {nombre_archivo} — sin registros")
        continue
    
    # Resumen: pivot por linea_negocio y concepto
    df_resumen = pd.pivot_table(
        df_filtrado,
        values="monto",
        index=["linea_negocio", "concepto"],
        aggfunc=["sum", "count"],
        fill_value=0,
    )
    df_resumen.columns = ["Monto Total", "Cantidad"]
    df_resumen = df_resumen.sort_values("Monto Total", ascending=False)
    
    filepath = os.path.join(OUTPUT_DIR, nombre_archivo)
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df_filtrado.to_excel(writer, sheet_name="Detalle", index=False)
        df_resumen.to_excel(writer, sheet_name="Resumen")
    
    total_monto = df_filtrado["monto"].sum()
    print(f"  ✅ {nombre_archivo} — {len(df_filtrado)} filas — Total: ${total_monto:,.2f}")

# Registros sin clasificar
df_sin_clasificar = df_consolidado[df_consolidado["clasificacion"] == "Sin clasificar"]
if len(df_sin_clasificar) > 0:
    filepath = os.path.join(OUTPUT_DIR, "no_clasificados.xlsx")
    df_sin_clasificar.to_excel(filepath, index=False)
    print(f"  ⚠️ no_clasificados.xlsx — {len(df_sin_clasificar)} registros sin clasificar")

# ═══════════════════════════════════════════════════════════
# PASO 5: FORMATO GLOBAL
# ═══════════════════════════════════════════════════════════

print("\n" + "=" * 60)
print("📊 PASO 5: Generando Formato Global")
print("=" * 60)

# Filtrar solo clasificados
df_clasificado = df_consolidado[df_consolidado["clasificacion"] != "Sin clasificar"].copy()

# Orden de líneas de negocio para las columnas
LINEAS_ORDEN = ["Banca Comercial", "Banca Corporativa", "Seguros",
                "Tarjetas", "Corporativo"]

# Función para generar sección del formato global
def generar_seccion(df, clasificacion):
    """Genera un pivot para una clasificación específica."""
    df_seccion = df[df["clasificacion"] == clasificacion]
    if len(df_seccion) == 0:
        return pd.DataFrame()
    
    pivot = pd.pivot_table(
        df_seccion,
        values="monto",
        index="concepto",
        columns="linea_negocio",
        aggfunc="sum",
        fill_value=0,
    )
    
    # Reordenar columnas y agregar las que falten
    for linea in LINEAS_ORDEN:
        if linea not in pivot.columns:
            pivot[linea] = 0
    pivot = pivot[LINEAS_ORDEN]
    
    # Agregar columna TOTAL
    pivot["TOTAL"] = pivot.sum(axis=1)
    
    # Ordenar por TOTAL descendente
    pivot = pivot.sort_values("TOTAL", ascending=False)
    
    return pivot

# Generar cada sección
pivot_ingresos = generar_seccion(df_clasificado, "Ingreso")
pivot_gastos = generar_seccion(df_clasificado, "Gasto")
pivot_provisiones = generar_seccion(df_clasificado, "Provisión")

# Subtotales
def fila_subtotal(pivot, nombre):
    """Crea una fila de subtotal."""
    subtotal = pivot.sum(axis=0)
    subtotal.name = nombre
    return pd.DataFrame([subtotal])

sub_ingresos = fila_subtotal(pivot_ingresos, "TOTAL INGRESOS")
sub_gastos = fila_subtotal(pivot_gastos, "TOTAL GASTOS")
sub_provisiones = fila_subtotal(pivot_provisiones, "TOTAL PROVISIONES")

# Resultado Neto
resultado_neto = sub_ingresos.iloc[0] - sub_gastos.iloc[0] - sub_provisiones.iloc[0]
resultado_neto.name = "RESULTADO NETO"
df_resultado = pd.DataFrame([resultado_neto])

# Concatenar todo
formato_global = pd.concat([
    pivot_ingresos,
    sub_ingresos,
    pd.DataFrame(index=[""]),  # Fila vacía separadora
    pivot_gastos,
    sub_gastos,
    pd.DataFrame(index=[""]),  # Fila vacía separadora
    pivot_provisiones,
    sub_provisiones,
    pd.DataFrame(index=[""]),  # Fila vacía separadora
    df_resultado,
])

# Guardar con formato
filepath_global = os.path.join(OUTPUT_DIR, "formato_global.xlsx")

with pd.ExcelWriter(filepath_global, engine="openpyxl") as writer:
    formato_global.to_excel(writer, sheet_name="Formato Global")
    
    # Aplicar formato
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    
    ws = writer.sheets["Formato Global"]
    
    # Formato de encabezados
    header_fill = PatternFill(start_color="0F62FE", end_color="0F62FE", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Formato de moneda y bordes
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    
    subtotal_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    resultado_fill = PatternFill(start_color="FFCD00", end_color="FFCD00", fill_type="solid")
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        concepto = row[0].value
        for cell in row:
            cell.border = thin_border
            if cell.column > 1 and isinstance(cell.value, (int, float)):
                cell.number_format = '$#,##0.00'
        
        # Formato especial para subtotales y resultado
        if concepto in ["TOTAL INGRESOS", "TOTAL GASTOS", "TOTAL PROVISIONES"]:
            for cell in row:
                cell.fill = subtotal_fill
                cell.font = Font(bold=True, size=11)
        elif concepto == "RESULTADO NETO":
            for cell in row:
                cell.fill = resultado_fill
                cell.font = Font(bold=True, size=12)
    
    # Auto-ajustar anchos
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 25)
    
    # Primera columna más ancha (conceptos)
    ws.column_dimensions["A"].width = 35

print(f"  ✅ formato_global.xlsx generado")
print(f"     {len(pivot_ingresos)} conceptos de ingreso")
print(f"     {len(pivot_gastos)} conceptos de gasto")
print(f"     {len(pivot_provisiones)} conceptos de provisión")

# ═══════════════════════════════════════════════════════════
# PASO 6: VALIDACIONES Y LOG DE EXCEPCIONES
# ═══════════════════════════════════════════════════════════

print("\n" + "=" * 60)
print("✅ PASO 6: Validaciones")
print("=" * 60)

excepciones = []

# 1. Montos nulos
nulos = df_consolidado[df_consolidado["monto"].isna()]
for _, row in nulos.iterrows():
    excepciones.append({
        "tipo_alerta": "Monto nulo",
        "archivo_origen": row.get("archivo_origen", ""),
        "concepto": row.get("concepto", ""),
        "monto": None,
        "detalle": "Registro con monto vacío o NaN",
    })
print(f"  {'✅' if len(nulos)==0 else '⚠️'} Montos nulos: {len(nulos)} registros")

# 2. Ingresos negativos
ingresos_neg = df_consolidado[
    (df_consolidado["clasificacion"] == "Ingreso") & 
    (df_consolidado["monto"] < 0)
]
for _, row in ingresos_neg.iterrows():
    excepciones.append({
        "tipo_alerta": "Ingreso negativo",
        "archivo_origen": row.get("archivo_origen", ""),
        "concepto": row.get("concepto", ""),
        "monto": row.get("monto", 0),
        "detalle": f"Ingreso con monto negativo: {row.get('monto', 0)}",
    })
print(f"  {'✅' if len(ingresos_neg)==0 else '⚠️'} Ingresos negativos: {len(ingresos_neg)} registros")

# 3. Cuadre de totales
total_entrada = len(df_consolidado)
total_ingresos = len(df_consolidado[df_consolidado["clasificacion"] == "Ingreso"])
total_gastos = len(df_consolidado[df_consolidado["clasificacion"] == "Gasto"])
total_provisiones = len(df_consolidado[df_consolidado["clasificacion"] == "Provisión"])
total_sin_clasificar = len(df_consolidado[df_consolidado["clasificacion"] == "Sin clasificar"])
total_salida = total_ingresos + total_gastos + total_provisiones + total_sin_clasificar

cuadre_ok = total_entrada == total_salida
print(f"  {'✅' if cuadre_ok else '❌'} Cuadre de totales: entrada={total_entrada}, salida={total_salida}")

# 4. Sin clasificar
print(f"  {'✅' if total_sin_clasificar==0 else '⚠️'} Sin clasificar: {total_sin_clasificar} registros")

# 5. Montos extremos (> 3 std del promedio)
df_montos_validos = df_consolidado.dropna(subset=["monto"])
media = df_montos_validos["monto"].mean()
std = df_montos_validos["monto"].std()
outliers = df_montos_validos[abs(df_montos_validos["monto"] - media) > 3 * std]
for _, row in outliers.iterrows():
    excepciones.append({
        "tipo_alerta": "Monto extremo",
        "archivo_origen": row.get("archivo_origen", ""),
        "concepto": row.get("concepto", ""),
        "monto": row.get("monto", 0),
        "detalle": f"Monto {row.get('monto', 0):,.2f} excede 3 desviaciones estándar (media={media:,.2f}, std={std:,.2f})",
    })
print(f"  {'✅' if len(outliers)==0 else '⚠️'} Montos extremos: {len(outliers)} registros")

# Guardar log de excepciones
if excepciones:
    df_excepciones = pd.DataFrame(excepciones)
    filepath_log = os.path.join(OUTPUT_DIR, "log_excepciones.xlsx")
    df_excepciones.to_excel(filepath_log, index=False)
    print(f"\n  📁 log_excepciones.xlsx — {len(excepciones)} alertas")

# ═══════════════════════════════════════════════════════════
# RESUMEN FINAL
# ═══════════════════════════════════════════════════════════

print("\n" + "═" * 60)
print("📊 RESUMEN DE CONSOLIDACIÓN")
print("═" * 60)
print(f"  Archivos leídos:          {total_archivos}")
print(f"  Registros totales:        {total_entrada:,}")
print(f"    → Ingresos:             {total_ingresos:,}")
print(f"    → Gastos:               {total_gastos:,}")
print(f"    → Provisiones:          {total_provisiones:,}")
print(f"    → Sin clasificar:       {total_sin_clasificar:,}  {'⚠️' if total_sin_clasificar > 0 else ''}")
print(f"  Alertas/excepciones:      {len(excepciones)}")
print(f"\n  Archivos generados en: {OUTPUT_DIR}")
for f in sorted(os.listdir(OUTPUT_DIR)):
    if f.endswith(('.xlsx', '.csv')):
        size_kb = os.path.getsize(os.path.join(OUTPUT_DIR, f)) / 1024
        print(f"    📁 {f} ({size_kb:.1f} KB)")
print("═" * 60)
print("✅ Consolidación completada exitosamente")
