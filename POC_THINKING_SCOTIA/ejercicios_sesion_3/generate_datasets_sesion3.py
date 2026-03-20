"""
Generador de datasets para Sesión 3 — POC Thinking: Múltiples fuentes → Formato Global

Genera 5 archivos Excel de entrada que simulan el flujo real de
consolidación financiera de una institución bancaria.

Fuentes generadas (590 filas total):
  1. banca_comercial.xlsx  — 150 filas (fecha_registro, monto_MXN, tipo_movimiento)
  2. banca_corporativa.xlsx — 120 filas (date, amount_MXN, type)
  3. seguros.xlsx          — 100 filas (fecha, importe, naturaleza)
  4. tarjetas.xlsx         — 130 filas (fecha_movimiento, monto_pesos, clasificacion)
  5. tesoreria.xlsx        —  90 filas (value_date, amount_MXN, pnl_type)

Cada archivo tiene columnas LIGERAMENTE DIFERENTES (simula realidad)
para que los participantes practiquen la homologación con Copilot 365.
"""

import random
import os
import csv
from datetime import datetime, timedelta

try:
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("❌ Necesitas instalar openpyxl: pip install openpyxl")
    exit(1)

random.seed(2026)

# Directorio de salida
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = os.path.join(SCRIPT_DIR, "datos_entrada")
os.makedirs(OUT_DIR, exist_ok=True)

# ═══════════════════════════════════════════════════════════
# DATOS COMPARTIDOS
# ═══════════════════════════════════════════════════════════

PERIODOS = ["2026-01", "2026-02", "2026-03"]
MESES_LABEL = {"2026-01": "Enero", "2026-02": "Febrero", "2026-03": "Marzo"}

LINEAS_NEGOCIO = [
    "Banca Comercial", "Banca Corporativa", "Seguros",
    "Inversiones", "Tarjetas", "Corporativo"
]

# Catálogo de conceptos con su clasificación y naturaleza
CONCEPTOS = {
    # Ingresos
    "Comisiones por apertura":       {"tipo": "Ingreso",    "cuenta": "4100", "subcuenta": "4110"},
    "Comisiones por manejo":         {"tipo": "Ingreso",    "cuenta": "4100", "subcuenta": "4120"},
    "Spread financiero":             {"tipo": "Ingreso",    "cuenta": "4200", "subcuenta": "4210"},
    "Intereses cobrados":            {"tipo": "Ingreso",    "cuenta": "4200", "subcuenta": "4220"},
    "Prima cobrada neta":            {"tipo": "Ingreso",    "cuenta": "4300", "subcuenta": "4310"},
    "Rendimiento de inversiones":    {"tipo": "Ingreso",    "cuenta": "4400", "subcuenta": "4410"},
    "Comisiones por cambio FX":      {"tipo": "Ingreso",    "cuenta": "4100", "subcuenta": "4130"},
    "Ingreso por asesoría":          {"tipo": "Ingreso",    "cuenta": "4500", "subcuenta": "4510"},
    "Ingreso por anualidad":         {"tipo": "Ingreso",    "cuenta": "4100", "subcuenta": "4140"},
    "Intereses tarjeta crédito":     {"tipo": "Ingreso",    "cuenta": "4200", "subcuenta": "4230"},
    # Gastos
    "Gastos de personal":            {"tipo": "Gasto",      "cuenta": "5100", "subcuenta": "5110"},
    "Nómina y prestaciones":         {"tipo": "Gasto",      "cuenta": "5100", "subcuenta": "5120"},
    "Servicios profesionales":       {"tipo": "Gasto",      "cuenta": "5200", "subcuenta": "5210"},
    "Licencias de software":         {"tipo": "Gasto",      "cuenta": "5300", "subcuenta": "5310"},
    "Renta de inmuebles":            {"tipo": "Gasto",      "cuenta": "5400", "subcuenta": "5410"},
    "Servicios de TI externos":      {"tipo": "Gasto",      "cuenta": "5300", "subcuenta": "5320"},
    "Mantenimiento y limpieza":      {"tipo": "Gasto",      "cuenta": "5500", "subcuenta": "5510"},
    "Multas regulatorias":           {"tipo": "Gasto",      "cuenta": "5600", "subcuenta": "5610"},
    "Gastos de publicidad":          {"tipo": "Gasto",      "cuenta": "5700", "subcuenta": "5710"},
    "Servicios de seguridad":        {"tipo": "Gasto",      "cuenta": "5500", "subcuenta": "5520"},
    # Provisiones
    "Reserva crediticia":            {"tipo": "Provisión",  "cuenta": "6100", "subcuenta": "6110"},
    "Pérdida esperada":              {"tipo": "Provisión",  "cuenta": "6100", "subcuenta": "6120"},
    "Reserva técnica seguros":       {"tipo": "Provisión",  "cuenta": "6200", "subcuenta": "6210"},
    "Provisión regulatoria":         {"tipo": "Provisión",  "cuenta": "6300", "subcuenta": "6310"},
    "Ajuste de valuación":           {"tipo": "Provisión",  "cuenta": "6400", "subcuenta": "6410"},
}

RESPONSABLES = [
    "Ana López", "Carlos Mendoza", "Sofía Hernández", "Miguel Ruiz",
    "Laura Castillo", "Patricia Díaz", "Javier Morales", "Fernanda Reyes",
    "Andrés García", "Daniela Vargas", "Ricardo Torres", "María Sánchez",
]


def random_date(periodo):
    """Genera una fecha aleatoria dentro del periodo YYYY-MM."""
    year, month = int(periodo[:4]), int(periodo[5:7])
    day = random.randint(1, 28)
    return datetime(year, month, day)


def format_money(val):
    """Formatea un número como moneda MXN."""
    return f"${val:,.2f}"


# ═══════════════════════════════════════════════════════════
# GENERADORES POR FUENTE
# ═══════════════════════════════════════════════════════════

def gen_linea_negocio(nombre_linea, conceptos_permitidos, n_rows, col_mapping):
    """
    Genera filas para una línea de negocio con nombres de columna específicos.
    col_mapping define cómo se llaman las columnas en este archivo particular.
    """
    rows = []
    for _ in range(n_rows):
        concepto_nombre = random.choice(conceptos_permitidos)
        concepto_info = CONCEPTOS[concepto_nombre]
        periodo = random.choice(PERIODOS)
        monto = round(random.uniform(50_000, 5_000_000), 2)
        # 3% de outliers
        if random.random() < 0.03:
            monto *= random.choice([3.5, 0.1, 5.0])
            monto = round(monto, 2)

        row = {
            col_mapping.get("fecha", "fecha"): random_date(periodo).strftime("%Y-%m-%d"),
            col_mapping.get("linea_negocio", "linea_negocio"): nombre_linea,
            col_mapping.get("concepto", "concepto"): concepto_nombre,
            col_mapping.get("subcuenta", "subcuenta"): concepto_info["subcuenta"],
            col_mapping.get("monto", "monto"): monto,
            col_mapping.get("tipo", "tipo"): concepto_info["tipo"],
            col_mapping.get("periodo", "periodo"): periodo,
            col_mapping.get("responsable", "responsable"): random.choice(RESPONSABLES),
        }
        rows.append(row)

    # Agregar datos sucios: 2% de filas con monto nulo o concepto vacío
    for i in random.sample(range(len(rows)), k=max(1, int(len(rows) * 0.02))):
        if random.random() < 0.5:
            rows[i][col_mapping.get("monto", "monto")] = None
        else:
            rows[i][col_mapping.get("concepto", "concepto")] = ""

    return rows


def write_xlsx(filename, rows, sheet_name="Datos"):
    """Escribe filas a un xlsx con formato de Tabla."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    if not rows:
        wb.save(os.path.join(OUT_DIR, filename))
        return

    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row[h] for h in headers])

    # Formato de Tabla
    table_ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"
    table_name = filename.replace(".xlsx", "").replace(" ", "_")
    tab = Table(displayName=table_name, ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Auto-ajustar anchos
    for col_idx, header in enumerate(headers, 1):
        max_len = max(len(str(header)), 12)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    filepath = os.path.join(OUT_DIR, filename)
    wb.save(filepath)
    print(f"  ✅ {filename} — {len(rows)} filas × {len(headers)} columnas")


def write_csv_file(filename, rows):
    """Escribe filas a un CSV."""
    if not rows:
        return
    filepath = os.path.join(OUT_DIR, filename)
    headers = list(rows[0].keys())
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)
    print(f"  ✅ {filename} — {len(rows)} filas × {len(headers)} columnas")


# ═══════════════════════════════════════════════════════════
# 1. BANCA COMERCIAL (columnas "estándar")
# ═══════════════════════════════════════════════════════════

def gen_banca_comercial():
    conceptos = [
        "Comisiones por apertura", "Comisiones por manejo", "Spread financiero",
        "Intereses cobrados", "Gastos de personal", "Servicios profesionales",
        "Reserva crediticia", "Pérdida esperada"
    ]
    col_map = {
        "fecha": "fecha_registro",
        "linea_negocio": "linea_negocio",
        "concepto": "concepto_operacion",
        "subcuenta": "subcuenta_contable",
        "monto": "monto_MXN",
        "tipo": "tipo_movimiento",
        "periodo": "periodo",
        "responsable": "responsable"
    }
    rows = gen_linea_negocio("Banca Comercial", conceptos, 150, col_map)
    write_xlsx("banca_comercial.xlsx", rows)


# ═══════════════════════════════════════════════════════════
# 2. BANCA CORPORATIVA (columnas diferentes)
# ═══════════════════════════════════════════════════════════

def gen_banca_corporativa():
    conceptos = [
        "Comisiones por apertura", "Spread financiero", "Intereses cobrados",
        "Ingreso por asesoría", "Gastos de personal", "Nómina y prestaciones",
        "Reserva crediticia", "Pérdida esperada"
    ]
    col_map = {
        "fecha": "date",
        "linea_negocio": "business_line",
        "concepto": "concept",
        "subcuenta": "sub_account",
        "monto": "amount_MXN",
        "tipo": "type",
        "periodo": "period",
        "responsable": "owner"
    }
    rows = gen_linea_negocio("Banca Corporativa", conceptos, 120, col_map)
    write_xlsx("banca_corporativa.xlsx", rows)


# ═══════════════════════════════════════════════════════════
# 3. SEGUROS
# ═══════════════════════════════════════════════════════════

def gen_seguros():
    conceptos = [
        "Prima cobrada neta", "Gastos de personal", "Servicios profesionales",
        "Reserva técnica seguros", "Provisión regulatoria", "Gastos de publicidad"
    ]
    col_map = {
        "fecha": "fecha",
        "linea_negocio": "area",
        "concepto": "tipo_operacion",
        "subcuenta": "cta_subcuenta",
        "monto": "importe",
        "tipo": "naturaleza",
        "periodo": "mes",
        "responsable": "responsable_area"
    }
    rows = gen_linea_negocio("Seguros", conceptos, 100, col_map)
    write_xlsx("seguros.xlsx", rows)


# ═══════════════════════════════════════════════════════════
# 4. TARJETAS
# ═══════════════════════════════════════════════════════════

def gen_tarjetas():
    conceptos = [
        "Comisiones por apertura", "Comisiones por manejo",
        "Intereses tarjeta crédito", "Ingreso por anualidad",
        "Gastos de personal", "Gastos de publicidad",
        "Reserva crediticia", "Pérdida esperada"
    ]
    col_map = {
        "fecha": "fecha_movimiento",
        "linea_negocio": "producto",
        "concepto": "concepto",
        "subcuenta": "num_cuenta",
        "monto": "monto_pesos",
        "tipo": "clasificacion",
        "periodo": "periodo_contable",
        "responsable": "ejecutivo"
    }
    rows = gen_linea_negocio("Tarjetas", conceptos, 130, col_map)
    write_xlsx("tarjetas.xlsx", rows)


# ═══════════════════════════════════════════════════════════
# 5. TESORERÍA
# ═══════════════════════════════════════════════════════════

def gen_tesoreria():
    conceptos = [
        "Comisiones por cambio FX", "Spread financiero",
        "Gastos de personal", "Servicios profesionales",
        "Ajuste de valuación"
    ]
    col_map = {
        "fecha": "value_date",
        "linea_negocio": "treasury_desk",
        "concepto": "transaction_type",
        "subcuenta": "account_code",
        "monto": "amount_MXN",
        "tipo": "pnl_type",
        "periodo": "month",
        "responsable": "dealer"
    }
    rows = gen_linea_negocio("Corporativo", conceptos, 90, col_map)
    # Tesorería reporta bajo "Corporativo"
    write_xlsx("tesoreria.xlsx", rows)


# ═══════════════════════════════════════════════════════════
# EJECUCIÓN
# ═══════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("=" * 60)
    print("🏗️  Generando datasets para Sesión 3")
    print(f"📁 Carpeta de salida: {OUT_DIR}")
    print("=" * 60)
    print()

    print("📊 Generando 5 archivos Excel:")
    gen_banca_comercial()
    gen_banca_corporativa()
    gen_seguros()
    gen_tarjetas()
    gen_tesoreria()

    print()
    print("=" * 60)
    total_files = len([f for f in os.listdir(OUT_DIR) if f.endswith(('.xlsx', '.csv'))])
    print(f"✅ {total_files} archivos generados en {OUT_DIR}")
    print("=" * 60)
